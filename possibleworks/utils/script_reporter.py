"""
Shared run-report helper for Server Scripts.

WHY THIS EXISTS
---------------
Server Scripts run inside safe_exec: no imports, so they cannot build a
spreadsheet, and duplicating ~100 lines of email HTML into every script means
13 places to change when the format changes.

This module is ordinary app code, so it can do what a Server Script cannot:
build an xlsx with openpyxl, render HTML tables, and attach files to email.
A Server Script calls it once at the end of its run:

    frappe.call(
        "possibleworks.utils.script_reporter.send_run_report",
        script_name="Monthly CL Allocation (GVS)",
        company="Ganges Valley School",
        summary={"Successful": 12, "Skipped": 3, "Failed": 1},
        rows=[
            {"employee": "GV001", "status": "Success", "action": "Created",
             "leaves": 1, "error": ""},
            {"employee": "GV009", "status": "Failed", "action": "",
             "leaves": "", "error": "Leave Type not found"},
        ],
        errors=["No active leave period for company"],   # run-level only
        recipients=["hr@example.com", "ops@example.com"],  # required
        dry_run=0,
    )

PER-EMPLOYEE ERRORS GO IN `rows`
-------------------------------
Give every record a row, including the ones that failed, with a `status` and an
`error` key. That way the spreadsheet shows the failure next to the employee it
belongs to and can be filtered on `status`. The separate `errors` argument is
for run-level problems that are not tied to any one employee.

frappe.call passes kwargs straight through after signature matching
(frappe/__init__.py:1117 -> get_newargs -> fn(**newargs)), so `rows` arrives as
a real list of dicts rather than a serialised string.

RECIPIENTS
----------
The calling script passes its own `recipients` list. There is no default and
no site-config fallback: a report goes only where the caller says it should.
With no usable address the report is skipped, not sent somewhere arbitrary.

Because a whitelisted method is reachable over HTTP by any logged-in session,
the call is gated two ways:

  1. _is_permitted() -- only Administrator or a System Manager / HR Manager /
     HR User may trigger a report at all. Anyone else is refused silently.
  2. _resolve_recipients() -- addresses are validated, de-duplicated and capped
     at MAX_RECIPIENTS, so the call cannot be turned into a bulk mailer.
"""

import io

import frappe

# Roles permitted to trigger a report. The scheduler runs as Administrator, so
# scheduled scripts always pass. A non-privileged caller does not raise -- the
# report is skipped so a calling script is never broken by this helper.
ALLOWED_ROLES = ("System Manager", "HR Manager", "HR User")

# Upper bound on how many addresses one report may go to.
MAX_RECIPIENTS = 20

# Rows shown inline in the email body. The xlsx attachment always carries all
# of them, so nothing is lost when this truncates.
ROW_PREVIEW_LIMIT = 50

# Hard ceiling on rows written to the attachment, to keep one runaway script
# from generating a 50MB email.
MAX_XLSX_ROWS = 20000


# =============================================================================
# PUBLIC ENTRY POINT
# =============================================================================

@frappe.whitelist()
def send_run_report(
	script_name,
	company=None,
	summary=None,
	rows=None,
	errors=None,
	recipients=None,
	dry_run=0,
	attach_xlsx=1,
):
	"""Email a single run report for a Server Script.

	Args:
		script_name: human name of the script, used in the subject.
		company:     optional company the run was scoped to.
		summary:     dict of counter name -> value.
		rows:        list of dicts, one per affected record. Keys become columns.
		errors:      list of strings (or dicts) describing per-record failures.
		recipients:  required list of email addresses. Validated, de-duplicated
		             and capped at MAX_RECIPIENTS. There is no default -- with
		             none supplied the report is skipped.
		dry_run:     truthy marks the subject and body as a dry run.
		attach_xlsx: truthy attaches the full row set as .xlsx.

	Returns:
		dict describing what happened. Never raises -- a reporting failure must
		not take down the script that called it.
	"""
	try:
		if not _is_permitted():
			frappe.logger().warning(
				f"script_reporter: refused report for '{script_name}' -- "
				f"user {frappe.session.user} lacks {ALLOWED_ROLES}"
			)
			return {"sent": False, "reason": "not permitted"}

		summary = _as_dict(summary)
		rows = _as_list(rows)
		errors = _as_list(errors)

		recipients = _resolve_recipients(recipients)

		if not recipients:
			frappe.logger().warning(
				f"script_reporter: no valid recipients passed, skipping report for '{script_name}'"
			)
			return {"sent": False, "reason": "no recipients"}

		is_dry_run = bool(frappe.utils.cint(dry_run))

		subject = _build_subject(script_name, company, summary, rows, errors, is_dry_run)
		message = _build_html(script_name, company, summary, rows, errors, is_dry_run)

		attachments = []

		if frappe.utils.cint(attach_xlsx) and rows:
			xlsx = _build_xlsx(script_name, rows)
			if xlsx:
				attachments.append({
					"fname": _safe_filename(script_name) + ".xlsx",
					"fcontent": xlsx,
				})

		frappe.sendmail(
			recipients=recipients,
			subject=subject,
			message=message,
			attachments=attachments or None,
		)

		# Errors also go to Error Log so they are searchable in the UI and
		# survive an email delivery failure. Per-row failures are included,
		# since those are the ones tied to a specific employee.
		row_failures = [
			f"{r.get('employee') or r.get('name') or '(unknown)'} | {r.get('error')}"
			for r in rows if _row_failed(r)
		]

		if errors or row_failures:
			frappe.log_error(
				title=f"{script_name} - {len(errors) + len(row_failures)} error(s) during run",
				message=_errors_as_text(script_name, company, list(errors) + row_failures),
			)

		return {
			"sent": True,
			"recipients": recipients,
			"rows": len(rows),
			"errors": len(errors),
			"failed_rows": len([r for r in rows if _row_failed(r)]),
			"attached": bool(attachments),
		}

	except Exception:
		# Swallow and log: the caller's own work is already done and must not be
		# rolled back or aborted because reporting failed.
		frappe.log_error(
			title=f"script_reporter failed for {script_name}",
			message=frappe.get_traceback(with_context=True),
		)
		return {"sent": False, "reason": "reporter raised"}


# =============================================================================
# PERMISSION / RECIPIENTS
# =============================================================================

def _is_permitted():
	if frappe.session.user == "Administrator":
		return True

	return bool(set(ALLOWED_ROLES) & set(frappe.get_roles()))


def _resolve_recipients(supplied=None):
	"""Caller-supplied addresses only -- there is no fallback list.

	The list is validated, de-duplicated and capped so this whitelisted method
	cannot be used as a bulk mailer. An empty result means the caller passed
	nothing usable, and send_run_report skips the report.
	"""
	valid = []

	candidates = _as_list(supplied)

	for address in candidates:
		address = str(address or "").strip()

		if not address or "@" not in address:
			continue

		if address in valid:
			continue

		valid.append(address)

		if len(valid) >= MAX_RECIPIENTS:
			frappe.logger().warning(
				f"script_reporter: recipient list capped at {MAX_RECIPIENTS}"
			)
			break

	return valid


# =============================================================================
# NORMALISATION
# =============================================================================

def _as_list(value):
	if value is None:
		return []
	if isinstance(value, (list, tuple)):
		return list(value)
	return [value]


def _as_dict(value):
	if value is None:
		return {}
	if isinstance(value, dict):
		return dict(value)
	return {"value": value}


def _cell(value):
	"""openpyxl and HTML both want primitives."""
	if value is None:
		return ""
	if isinstance(value, (str, int, float, bool)):
		return value
	return str(value)


def _row_columns(rows):
	"""Union of keys across all rows, preserving first-seen order.

	`error` is forced to the last column so a wide row still ends with the
	failure reason, which is what someone scanning the sheet looks for.
	"""
	columns = []
	for row in rows:
		if not isinstance(row, dict):
			continue
		for key in row.keys():
			if key not in columns:
				columns.append(key)

	if "error" in columns:
		columns.remove("error")
		columns.append("error")

	return columns


def _row_failed(row):
	"""A row counts as failed if it carries an error, or says so in `status`."""
	if not isinstance(row, dict):
		return False

	if str(row.get("error") or "").strip():
		return True

	return str(row.get("status") or "").strip().lower() in ("failed", "error")


# =============================================================================
# SUBJECT / BODY
# =============================================================================

def _build_subject(script_name, company, summary, rows, errors, is_dry_run):
	parts = []

	if is_dry_run:
		parts.append("[DRY RUN]")

	failed_count = len(errors) + len([r for r in _as_list(rows) if _row_failed(r)])

	if failed_count:
		parts.append(f"[{failed_count} ERROR(S)]")

	parts.append(script_name)

	if company:
		parts.append(f"- {company}")

	counters = " | ".join(f"{k}: {summary[k]}" for k in summary)
	if counters:
		parts.append(f"({counters})")

	return " ".join(parts)


def _build_html(script_name, company, summary, rows, errors, is_dry_run):
	banner = ""

	if is_dry_run:
		banner = (
			'<div style="background:#e8f4fd;padding:12px;border-left:4px solid #1a73e8;margin:16px 0">'
			"<strong>DRY RUN</strong> - nothing was written."
			"</div>"
		)
	elif errors:
		banner = (
			'<div style="background:#fff3e0;padding:12px;border-left:4px solid #f57c00;margin:16px 0">'
			f"<strong>{len(errors)} record(s) failed.</strong> Details below and in Error Log."
			"</div>"
		)

	meta_rows = [
		("Script", script_name),
		("Company", company or "(all in scope)"),
		("Site", frappe.local.site),
		("Run at", frappe.utils.now()),
		("Triggered by", frappe.session.user),
	]

	meta_html = "".join(
		'<tr><td style="padding:6px 12px;border:1px solid #ddd;background:#f5f5f5;width:150px">'
		f"<strong>{label}</strong></td>"
		f'<td style="padding:6px 12px;border:1px solid #ddd">{frappe.utils.escape_html(str(value))}</td></tr>'
		for label, value in meta_rows
	)

	summary_html = ""
	if summary:
		cells = "".join(
			'<tr><td style="padding:6px 12px;border:1px solid #ddd;background:#f5f5f5">'
			f"<strong>{frappe.utils.escape_html(str(k))}</strong></td>"
			f'<td style="padding:6px 12px;border:1px solid #ddd">{frappe.utils.escape_html(str(summary[k]))}</td></tr>'
			for k in summary
		)
		summary_html = (
			"<h3 style=\"margin-top:24px\">Summary</h3>"
			'<table style="border-collapse:collapse;font-size:14px">' + cells + "</table>"
		)

	rows_html = ""
	if rows:
		columns = _row_columns(rows)

		# Failed rows first, so a truncated preview never hides them.
		failed_rows = [r for r in rows if _row_failed(r)]
		ok_rows = [r for r in rows if not _row_failed(r)]
		ordered = failed_rows + ok_rows
		shown = ordered[:ROW_PREVIEW_LIMIT]

		head = "".join(
			'<th style="padding:8px;border:1px solid #ddd;background:#f5f7fb;text-align:left">'
			f"{frappe.utils.escape_html(str(c))}</th>"
			for c in columns
		)

		body = ""
		for row in shown:
			if not isinstance(row, dict):
				continue
			tint = ' style="background:#ffe5e5"' if _row_failed(row) else ""
			body += f"<tr{tint}>" + "".join(
				'<td style="padding:8px;border:1px solid #ddd">'
				f"{frappe.utils.escape_html(str(_cell(row.get(c))))}</td>"
				for c in columns
			) + "</tr>"

		notes = []

		if failed_rows:
			notes.append(
				f'<p style="color:#d32f2f;font-size:13px"><strong>{len(failed_rows)} '
				f"record(s) failed</strong> - listed first below and in the spreadsheet.</p>"
			)

		if len(ordered) > ROW_PREVIEW_LIMIT:
			notes.append(
				f'<p style="color:#666;font-size:13px">Showing first {ROW_PREVIEW_LIMIT} '
				f"of {len(rows)} rows. The attached spreadsheet has all of them.</p>"
			)

		rows_html = (
			f'<h3 style="margin-top:24px">Records ({len(rows)})</h3>'
			+ "".join(notes)
			+ '<table style="border-collapse:collapse;font-size:13px;width:100%">'
			+ "<thead><tr>" + head + "</tr></thead><tbody>" + body + "</tbody></table>"
		)

	errors_html = ""
	if errors:
		items = "".join(
			f'<li style="margin-bottom:4px">{frappe.utils.escape_html(str(e))}</li>'
			for e in errors
		)
		errors_html = (
			f'<h3 style="margin-top:24px;color:#d32f2f">Errors ({len(errors)})</h3>'
			f'<ul style="font-size:13px;color:#3d475c">{items}</ul>'
		)

	return f"""
	<div style="font-family:Arial,sans-serif;max-width:900px">
		<h2 style="color:#0f1e3d">{frappe.utils.escape_html(script_name)}</h2>
		{banner}
		<table style="border-collapse:collapse;font-size:14px">{meta_html}</table>
		{summary_html}
		{rows_html}
		{errors_html}
		<p style="color:#666;font-size:12px;margin-top:28px">
			Generated by possibleworks.utils.script_reporter. Full console output
			is in the site log; failures are also in Error Log.
		</p>
	</div>
	"""


# =============================================================================
# XLSX
# =============================================================================

def _build_xlsx(script_name, rows):
	"""Return xlsx bytes, or None if it could not be built."""
	try:
		from openpyxl import Workbook
		from openpyxl.styles import Font

		columns = _row_columns(rows)

		if not columns:
			return None

		wb = Workbook()
		ws = wb.active
		ws.title = "Run Report"

		ws.append([str(c) for c in columns])

		for cell in ws[1]:
			cell.font = Font(bold=True)

		# Failed rows first so they are visible without scrolling or filtering.
		ordered = [r for r in rows if _row_failed(r)] + [r for r in rows if not _row_failed(r)]
		capped = ordered[:MAX_XLSX_ROWS]

		for row in capped:
			if not isinstance(row, dict):
				continue
			ws.append([_cell(row.get(c)) for c in columns])

		if len(rows) > MAX_XLSX_ROWS:
			ws.append([])
			ws.append([f"TRUNCATED: {len(rows)} rows produced, first {MAX_XLSX_ROWS} written"])

		# Readable column widths without measuring every cell.
		for idx, column in enumerate(columns, start=1):
			width = max(12, min(40, len(str(column)) + 6))
			ws.column_dimensions[ws.cell(row=1, column=idx).column_letter].width = width

		ws.freeze_panes = "A2"

		buffer = io.BytesIO()
		wb.save(buffer)
		return buffer.getvalue()

	except Exception:
		frappe.log_error(
			title=f"script_reporter: xlsx build failed for {script_name}",
			message=frappe.get_traceback(with_context=True),
		)
		return None


# =============================================================================
# HELPERS
# =============================================================================

def _safe_filename(script_name):
	keep = []
	for ch in str(script_name):
		keep.append(ch if (ch.isalnum() or ch in ("-", "_")) else "_")
	return ("".join(keep) or "run_report")[:80]


def _errors_as_text(script_name, company, errors):
	lines = [
		f"Script: {script_name}",
		f"Company: {company or '(all in scope)'}",
		f"Site: {frappe.local.site}",
		f"Errors: {len(errors)}",
		"",
	]
	for e in errors:
		lines.append(f"  {e}")
	return "\n".join(lines)
